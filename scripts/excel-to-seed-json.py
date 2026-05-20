#!/usr/bin/env python3
"""
Excel to Seed JSON converter for Trazasole project.
Reads Trazasole_Datos_Consolidados.xlsx and outputs seed JSON files.
"""

import json
import os
import re
import sys
from datetime import datetime, date

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------
EXCEL_PATH = "/home/z/my-project/trz7/Trazasole_Datos_Consolidados.xlsx"
if not os.path.exists(EXCEL_PATH):
    EXCEL_PATH = "/home/z/my-project/download/Trazasole_Datos_Consolidados.xlsx"

OUTPUT_DIR = "/home/z/my-project/trz7/prisma/seed-data"

os.makedirs(OUTPUT_DIR, exist_ok=True)

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def fmt_date(val):
    """Convert a date/datetime to ISO string or return None."""
    if val is None:
        return None
    if isinstance(val, (datetime, date)):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, str):
        val = val.strip()
        if not val:
            return None
        return val
    return None


def fmt_num(val):
    """Convert to int or float, or None."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        if isinstance(val, float) and val == int(val):
            return int(val)
        return val
    if isinstance(val, str):
        val = val.strip()
        if not val:
            return None
        try:
            f = float(val)
            return int(f) if f == int(f) else f
        except ValueError:
            return None
    return None


def clean_str(val):
    """Return stripped string or None."""
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


def read_sheet(wb, sheet_name):
    """Read all rows (skip header) as list of dicts using header row."""
    ws = wb[sheet_name]
    headers = []
    for cell in next(ws.iter_rows(min_row=1, max_row=1)):
        v = cell.value
        headers.append(str(v).strip() if v else None)
    
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rows.append(row)
    return headers, rows


def row_dict(headers, row):
    """Convert a row tuple to a dict using headers."""
    d = {}
    for i, h in enumerate(headers):
        if h is not None:
            d[h] = row[i] if i < len(row) else None
    return d


def save_json(data, filename):
    path = os.path.join(OUTPUT_DIR, filename)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print(f"  -> {path} ({len(data)} records)")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    print(f"Reading Excel: {EXCEL_PATH}")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    print(f"Sheets: {wb.sheetnames}\n")

    # ===================================================================
    # 1. CLIENTES (Titulares)
    # ===================================================================
    print("=== Processing Titulares -> clientes.json ===")
    headers, rows = read_sheet(wb, "Titulares")

    # First pass: collect productor and usuario faena names from Tropas
    tropa_headers, tropa_rows = read_sheet(wb, "Tropas")
    productor_names = set()
    usuario_faena_names = set()
    for row in tropa_rows:
        d = row_dict(tropa_headers, row)
        prod = clean_str(d.get("Productor"))
        uf = clean_str(d.get("Matarife / Usuario Faena"))
        if prod:
            productor_names.add(prod.upper())
        if uf:
            usuario_faena_names.add(uf.upper())

    def normalize_name(name):
        """Normalize a name for fuzzy matching: remove suffixes, extra spaces, dots."""
        if not name:
            return ""
        n = name.upper()
        # Remove common legal suffixes
        for suffix in [" S.R.L.", " SRL", " S.A.S", " SAS", " S.A.", " SA", " S.A", " S.A.S."]:
            n = n.replace(suffix, "")
        # Remove trailing dots, commas, and extra whitespace
        n = re.sub(r"[.,\s]+", " ", n).strip()
        return n

    def stem_word(w, min_len=4):
        """Simple stemming: strip common Spanish plural/conjugation endings."""
        if len(w) <= min_len:
            return w
        for suffix in ["OS", "AS", "ES", "ONES"]:
            if w.endswith(suffix) and len(w) - len(suffix) >= min_len:
                return w[:-len(suffix)]
        return w

    def prefix_match(a, b, min_prefix=4):
        """Check if two words share a common prefix of min_prefix length."""
        if len(a) < min_prefix or len(b) < min_prefix:
            return False
        return a[:min_prefix] == b[:min_prefix]

    def names_match(a, b):
        """Check if two names likely refer to the same entity."""
        na, nb = normalize_name(a), normalize_name(b)
        if na == nb:
            return True
        words_a = set(w for w in na.split() if len(w) > 2)
        words_b = set(w for w in nb.split() if len(w) > 2)
        if words_a and words_b:
            overlap = words_a & words_b
            if len(overlap) >= 2:
                return True
            if len(overlap) >= 1 and (words_a <= words_b or words_b <= words_a):
                return True
            # Prefix matching for near-matches like BOSQUES/BOSQUE, AMADOS/AMADO
            prefix_matches = sum(
                1 for wa in words_a for wb in words_b if prefix_match(wa, wb)
            )
            if prefix_matches >= 2:
                return True
        return False

    clientes = []
    for idx, row in enumerate(rows):
        d = row_dict(headers, row)
        nombre = clean_str(d.get("Nombre"))
        if not nombre:
            continue

        nombre_upper = nombre.upper()
        es_productor = any(names_match(nombre_upper, pn) for pn in productor_names)
        es_usuario_faena = any(names_match(nombre_upper, un) for un in usuario_faena_names)

        clientes.append({
            "id": f"cli-{idx+1:03d}",
            "nombre": nombre,
            "cuit": clean_str(d.get("CUIT")),
            "email": clean_str(d.get("Mail")),
            "telefono": clean_str(d.get("Celular")),
            "contacto": clean_str(d.get("Contacto")),
            "esProductor": es_productor,
            "esUsuarioFaena": es_usuario_faena,
        })

    save_json(clientes, "clientes.json")

    # ===================================================================
    # 2. CORRALES
    # ===================================================================
    print("\n=== Processing Corrales -> corrales.json ===")
    headers, rows = read_sheet(wb, "Corrales")
    corrales = []
    for idx, row in enumerate(rows):
        d = row_dict(headers, row)
        nombre = clean_str(d.get("Nombre"))
        if not nombre:
            continue
        corrales.append({
            "id": f"corral-{idx+1:02d}",
            "nombre": nombre,
            "capacidad": fmt_num(d.get("Capacidad")),
            "observaciones": clean_str(d.get("Observaciones")),
        })

    save_json(corrales, "corrales.json")

    # ===================================================================
    # 3. CAMARAS
    # ===================================================================
    print("\n=== Processing Camaras -> camaras.json ===")
    headers, rows = read_sheet(wb, "Camaras")
    camaras = []
    for idx, row in enumerate(rows):
        d = row_dict(headers, row)
        nombre = clean_str(d.get("Nombre"))
        if not nombre:
            continue

        tipo_raw = clean_str(d.get("Tipo")) or ""
        tipo_upper = tipo_raw.upper()
        if "CUARTEO" in tipo_upper:
            tipo = "CUARTEO"
        elif "FAENA" in tipo_upper:
            tipo = "FAENA"
        else:
            tipo = "DEPOSITO"

        camaras.append({
            "id": f"camara-{idx+1:02d}",
            "nombre": nombre,
            "tipo": tipo,
            "capacidad": fmt_num(d.get("Capacidad")),
            "observaciones": clean_str(d.get("Observaciones")),
        })

    save_json(camaras, "camaras.json")

    # ===================================================================
    # 4. TIPIFICADORES
    # ===================================================================
    print("\n=== Processing Tipificadores -> tipificadores.json ===")
    headers, rows = read_sheet(wb, "Tipificadores")
    tipificadores = []
    for idx, row in enumerate(rows):
        d = row_dict(headers, row)
        nombre = clean_str(d.get("Nombre"))
        if not nombre:
            continue
        tipificadores.append({
            "id": f"tip-{idx+1:02d}",
            "nombre": nombre,
            "apellido": clean_str(d.get("Apellido")),
            "numero": clean_str(d.get("Nº Interno")),
            "matricula": clean_str(d.get("Matrícula")),
        })

    save_json(tipificadores, "tipificadores.json")

    # ===================================================================
    # 5. TROPAS
    # ===================================================================
    print("\n=== Processing Tropas -> tropas.json ===")
    headers, rows = read_sheet(wb, "Tropas")
    
    # Build a tropa fecha lookup for romaneos later
    tropa_fecha_faena = {}  # tropa_numero -> fecha faena ISO string
    tropa_fecha_ingreso = {}

    tropas = []
    for row in rows:
        d = row_dict(headers, row)
        numero = fmt_num(d.get("No Tropa"))
        if numero is None:
            continue

        fecha_faena = fmt_date(d.get("Fecha Faena"))
        fecha_ingreso = fmt_date(d.get("Fecha Ingreso"))

        # Skip tropas without Fecha Faena (empty tropas 173-200 and Ingresada stubs)
        if fecha_faena is None:
            continue

        # Map estado
        estado_raw = clean_str(d.get("Estado")) or ""
        estado_map = {
            "FAENADA": "FAENADO",
            "INGRESADA": "RECIBIDO",
        }
        estado = estado_map.get(estado_raw.upper(), estado_raw.upper())

        tropa_codigo = clean_str(d.get("Codigo"))
        
        # Store fecha lookups
        tropa_fecha_faena[numero] = fecha_faena
        tropa_fecha_ingreso[numero] = fecha_ingreso

        # Resolve usuario de faena: prioridad Matarife > CUIT Titular > CUIT Matarife
        uf_nombre = clean_str(d.get("Matarife / Usuario Faena"))
        uf_cuit = clean_str(d.get("CUIT Matarife"))
        titular_cuit = clean_str(d.get("CUIT Titular"))
        titular_contacto = clean_str(d.get("Contacto"))

        # Si no tiene matarife, buscar por CUIT del titular
        if not uf_nombre and titular_cuit:
            # Buscar en la lista de titulares por CUIT
            for cl in clientes:
                if cl.get("cuit") and cl["cuit"].upper() == titular_cuit.upper():
                    uf_nombre = cl["nombre"]
                    break

        # Si no tiene CUIT Productor, buscar por CUIT Titular (a veces coincide)
        prod_nombre = clean_str(d.get("Productor"))
        prod_cuit = clean_str(d.get("CUIT Productor"))

        tropas.append({
            "numero": numero,
            "codigo": tropa_codigo,
            "especie": clean_str(d.get("Especie")),
            "cantidadCabezas": fmt_num(d.get("Cant. Cabezas")),
            "productorNombre": prod_nombre,
            "productorCuit": prod_cuit,
            "usuarioFaenaNombre": uf_nombre,
            "usuarioFaenaCuit": uf_cuit,
            "dte": clean_str(d.get("DTE")),
            "guia": clean_str(d.get("Guia")),
            "corral": clean_str(d.get("Corral")),
            "matriculaMatarife": clean_str(d.get("Matricula Matarife")),
            "cuitTitular": titular_cuit,
            "contactoTitular": titular_contacto,
            "fechaFaena": fecha_faena,
            "fechaIngreso": fecha_ingreso,
            "estado": estado,
            "pesoVivo": fmt_num(d.get("Kg Vivo")),
            "kgGancho": fmt_num(d.get("Kg 1/2 Res")),
            "rindePct": fmt_num(d.get("Rinde %")),
            "observaciones": clean_str(d.get("Observaciones Faena")),
        })

    save_json(tropas, "tropas.json")

    # ===================================================================
    # 6. ANIMALES
    # ===================================================================
    print("\n=== Processing Animales -> animales.json ===")
    headers, rows = read_sheet(wb, "Animales")
    
    animales = []
    for row in rows:
        d = row_dict(headers, row)
        tropa_codigo = clean_str(d.get("Código Tropa"))
        tropa_numero = fmt_num(d.get("Nº Tropa"))
        animal_numero = fmt_num(d.get("Nº Animal"))
        
        if tropa_numero is None:
            continue

        # Generate codigo: "B 2026 0001" -> "B20260001", then "B20260001-001"
        codigo_base = (tropa_codigo or "").replace(" ", "")
        codigo = f"{codigo_base}-{animal_numero:03d}" if animal_numero else None

        # Map estado
        estado_raw = clean_str(d.get("Estado")) or ""
        if "rinde" in estado_raw.lower():
            estado = "FAENADO"
        else:
            estado = estado_raw.upper()

        animales.append({
            "tropaNumero": tropa_numero,
            "numero": animal_numero,
            "codigo": codigo,
            "caravana": clean_str(d.get("Caravana")),
            "raza": clean_str(d.get("Raza")),
            "tipoAnimal": clean_str(d.get("Tipo Animal")),
            "pesoVivo": fmt_num(d.get("Peso Vivo (ingreso)")),
            "kgEntrada": fmt_num(d.get("Kg Entrada (rinde)")),
            "kgMediaA": fmt_num(d.get("Kg Media A")),
            "kgMediaB": fmt_num(d.get("Kg Media B")),
            "totalKg": fmt_num(d.get("Total Kg")),
            "rindePct": fmt_num(d.get("Rinde %")),
            "garron": fmt_num(d.get("Garrón")),
            "estado": estado,
        })

    save_json(animales, "animales.json")

    # ===================================================================
    # 7. FACTURAS
    # ===================================================================
    print("\n=== Processing Facturacion -> facturas.json ===")
    headers, rows = read_sheet(wb, "Facturacion")
    
    facturas = []
    for row in rows:
        d = row_dict(headers, row)
        tropa_numero = fmt_num(d.get("No Tropa"))
        if tropa_numero is None:
            continue

        # Map estado pago
        estado_pago_raw = clean_str(d.get("Estado Pago")) or ""
        estado_pago_upper = estado_pago_raw.upper()
        if estado_pago_upper == "COBRADO":
            estado_pago = "PAGADO"
        elif estado_pago_raw and not any(c.isalpha() for c in estado_pago_raw):
            # Numeric value - treat as partial indicator
            estado_pago = "PENDIENTE"
        else:
            estado_pago = "PENDIENTE"

        facturas.append({
            "tropaNumero": tropa_numero,
            "usuario": clean_str(d.get("Usuario")),
            "fechaFaena": fmt_date(d.get("Fecha Faena")),
            "cantAnimales": fmt_num(d.get("Cant. Animales")),
            "kgGancho": fmt_num(d.get("Kg Gancho")),
            "totalServicioIVA": fmt_num(d.get("Total Servicio +IVA")),
            "tasaInspVet": fmt_num(d.get("Tasa Insp. Vet")),
            "arancelIPCVA": fmt_num(d.get("Arancel IPCVA")),
            "totalFacturaImp": fmt_num(d.get("Total Factura c/Imp")),
            "noFactura": clean_str(d.get("No Factura")),
            "fechaFactura": fmt_date(d.get("Fecha Factura")),
            "fechaPago": fmt_date(d.get("Fecha Pago")),
            "montoDepositado": fmt_num(d.get("Monto Depositado")),
            "estadoPago": estado_pago,
        })

    save_json(facturas, "facturas.json")

    # ===================================================================
    # 8. ROMANEOS (Animales with kgEntrada > 0)
    # ===================================================================
    print("\n=== Processing Animales (rinde) -> romaneos.json ===")
    headers_anim, rows_anim = read_sheet(wb, "Animales")
    
    # Build tropa codigo lookup: tropa_numero -> tropa_codigo
    tropa_codigo_lookup = {}
    for t in tropas:
        tropa_codigo_lookup[t["numero"]] = t["codigo"]

    romaneos = []
    for row in rows_anim:
        d = row_dict(headers_anim, row)
        tropa_numero = fmt_num(d.get("Nº Tropa"))
        animal_numero = fmt_num(d.get("Nº Animal"))
        kg_entrada = fmt_num(d.get("Kg Entrada (rinde)"))
        
        if tropa_numero is None or kg_entrada is None or kg_entrada <= 0:
            continue

        # Get fecha from tropa
        fecha_romaneo = tropa_fecha_faena.get(tropa_numero)
        if not fecha_romaneo:
            fecha_romaneo = tropa_fecha_ingreso.get(tropa_numero)
        if not fecha_romaneo:
            fecha_romaneo = datetime.now().strftime("%Y-%m-%d")

        tropa_codigo = tropa_codigo_lookup.get(tropa_numero, "")

        romaneos.append({
            "tropaNumero": tropa_numero,
            "animalNumero": animal_numero,
            "garron": fmt_num(d.get("Garrón")),
            "tropaCodigo": tropa_codigo,
            "tipoAnimal": clean_str(d.get("Tipo Animal")),
            "raza": clean_str(d.get("Raza")),
            "pesoVivo": fmt_num(d.get("Peso Vivo (ingreso)")),
            "kgMediaIzq": fmt_num(d.get("Kg Media A")),
            "kgMediaDer": fmt_num(d.get("Kg Media B")),
            "pesoTotal": fmt_num(d.get("Total Kg")),
            "rinde": fmt_num(d.get("Rinde %")),
            "fecha": fecha_romaneo,
            "estado": "CONFIRMADO",
        })

    save_json(romaneos, "romaneos.json")

    # ===================================================================
    # 9. MENUDENCIAS
    # ===================================================================
    print("\n=== Processing Menudencias -> menudencias.json ===")
    headers, rows = read_sheet(wb, "Menudencias")
    
    menudencias = []
    for row in rows:
        d = row_dict(headers, row)
        tropa_numero = fmt_num(d.get("No Tropa"))
        item = clean_str(d.get("Item"))
        if tropa_numero is None:
            continue

        menudencias.append({
            "tropaNumero": tropa_numero,
            "item": item,
            "cantidad": fmt_num(d.get("Cantidad")),
            "kg": fmt_num(d.get("Kg")),
            "unidades": fmt_num(d.get("Unidades")),
            "kgDecomiso": fmt_num(d.get("Kg Decomiso")),
            "tipoMenudenciaNombre": item,
        })

    save_json(menudencias, "menudencias.json")

    # ===================================================================
    # Summary
    # ===================================================================
    print("\n" + "=" * 60)
    print("SUMMARY STATISTICS")
    print("=" * 60)
    print(f"  clientes.json:       {len(clientes):>5} records")
    print(f"  corrales.json:       {len(corrales):>5} records")
    print(f"  camaras.json:        {len(camaras):>5} records")
    print(f"  tipificadores.json:  {len(tipificadores):>5} records")
    print(f"  tropas.json:         {len(tropas):>5} records")
    print(f"  animales.json:       {len(animales):>5} records")
    print(f"  facturas.json:       {len(facturas):>5} records")
    print(f"  romaneos.json:       {len(romaneos):>5} records")
    print(f"  menudencias.json:    {len(menudencias):>5} records")
    print(f"  TOTAL:               {len(clientes)+len(corrales)+len(camaras)+len(tipificadores)+len(tropas)+len(animales)+len(facturas)+len(romaneos)+len(menudencias):>5} records")
    print(f"\nOutput directory: {OUTPUT_DIR}")
    print("Done!")


if __name__ == "__main__":
    main()
