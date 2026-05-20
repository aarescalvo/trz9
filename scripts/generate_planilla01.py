#!/usr/bin/env python3
"""
Generador de Planilla 01 - SENASA
Genera un archivo Excel con el formato oficial de Planilla 01 para registro de ingreso de hacienda.
"""

import sys
import json
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

def get_week_number(date_str):
    """Obtiene el número de semana del año"""
    try:
        # Parsear fecha ISO y convertir a fecha naive
        d = datetime.fromisoformat(date_str.replace('Z', '+00:00'))
        d = d.replace(tzinfo=None)
    except:
        d = datetime.strptime(date_str.split('T')[0], '%Y-%m-%d')
    start = datetime(d.year, 1, 1)
    diff = d - start
    return (diff.days + start.weekday() + 7) // 7

def get_tipo_label(tipo):
    """Obtiene la etiqueta del tipo de animal"""
    tipos = {
        'TO': 'TORO',
        'VA': 'VACA',
        'VQ': 'VAQUILLONA',
        'MEJ': 'MEJ',
        'NO': 'NOVILLO',
        'NT': 'NOVILLITO',
    }
    return tipos.get(tipo, tipo)

def create_planilla_01(data, output_path):
    """Crea la Planilla 01 en formato Excel"""
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Planilla 01"
    
    # Estilos
    title_font = Font(name='Arial', size=14, bold=True)
    header_font = Font(name='Arial', size=10, bold=True)
    normal_font = Font(name='Arial', size=10)
    small_font = Font(name='Arial', size=8)
    
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center')
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    header_fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
    
    # Anchos de columna
    column_widths = [4, 15, 15, 15, 15, 15, 15, 15, 15, 15, 15, 15]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    # Fila 1: Título
    ws.merge_cells('A1:L1')
    ws['A1'] = 'PLANILLA 01 - BOVINO'
    ws['A1'].font = title_font
    ws['A1'].alignment = center_align
    
    # Fila 2: Encabezado empresa
    ws.merge_cells('A2:D2')
    ws['A2'] = 'Solemar Alimentaria S.A.'
    ws['A2'].font = header_font
    ws['A2'].alignment = left_align
    
    ws.merge_cells('E2:G2')
    ws['E2'] = f"N° SENASA: 3986 | Matrícula: 300"
    ws['E2'].font = normal_font
    ws['E2'].alignment = center_align
    
    ws.merge_cells('H2:L2')
    ws['H2'] = f"Semana N°: {get_week_number(data.get('fechaRecepcion', datetime.now().isoformat()))}"
    ws['H2'].font = normal_font
    ws['H2'].alignment = center_align
    
    # Fila 3: Espacio
    ws.row_dimensions[3].height = 10
    
    # Fila 4: Productor
    ws.merge_cells('A4:B4')
    ws['A4'] = 'Productor:'
    ws['A4'].font = header_font
    ws.merge_cells('C4:F4')
    ws['C4'] = data.get('productor', {}).get('nombre', '-')
    ws['C4'].font = normal_font
    
    ws.merge_cells('G4:H4')
    ws['G4'] = 'CUIT:'
    ws['G4'].font = header_font
    ws.merge_cells('I4:L4')
    ws['I4'] = data.get('productor', {}).get('cuit', '-')
    ws['I4'].font = normal_font
    
    # Fila 5: Usuario/Matarife
    ws.merge_cells('A5:B5')
    ws['A5'] = 'Usuario/Matarife:'
    ws['A5'].font = header_font
    ws.merge_cells('C5:F5')
    ws['C5'] = data.get('usuarioFaena', {}).get('nombre', '-')
    ws['C5'].font = normal_font
    
    ws.merge_cells('G5:H5')
    ws['G5'] = 'CUIT:'
    ws['G5'].font = header_font
    ws.merge_cells('I5:L5')
    ws['I5'] = data.get('usuarioFaena', {}).get('cuit', '-')
    ws['I5'].font = normal_font
    
    # Fila 6: Fecha y Tropa
    ws.merge_cells('A6:B6')
    ws['A6'] = 'Fecha Ingreso:'
    ws['A6'].font = header_font
    ws['C6'] = datetime.fromisoformat(data.get('fechaRecepcion', datetime.now().isoformat()).replace('Z', '+00:00')).strftime('%d/%m/%Y') if data.get('fechaRecepcion') else '-'
    ws['C6'].font = normal_font
    
    ws.merge_cells('D6:E6')
    ws['D6'] = 'N° Tropa:'
    ws['D6'].font = header_font
    ws.merge_cells('F6:G6')
    ws['F6'] = data.get('codigo', '-')
    ws['F6'].font = normal_font
    
    ws.merge_cells('H6:I6')
    ws['H6'] = 'Cantidad:'
    ws['H6'].font = header_font
    ws.merge_cells('J6:L6')
    ws['J6'] = f"{data.get('cantidadCabezas', 0)} cabezas"
    ws['J6'].font = normal_font
    
    # Fila 7: Transportista
    pesaje = data.get('pesajeCamion', {}) or {}
    transportista = pesaje.get('transportista', {}) or {}
    
    ws.merge_cells('A7:B7')
    ws['A7'] = 'Transportista:'
    ws['A7'].font = header_font
    ws.merge_cells('C7:F7')
    ws['C7'] = transportista.get('nombre', '-')
    ws['C7'].font = normal_font
    
    ws.merge_cells('G7:H7')
    ws['G7'] = 'CUIT:'
    ws['G7'].font = header_font
    ws.merge_cells('I7:L7')
    ws['I7'] = transportista.get('cuit', '-')
    ws['I7'].font = normal_font
    
    # Fila 8: Chofer
    ws.merge_cells('A8:B8')
    ws['A8'] = 'Chofer:'
    ws['A8'].font = header_font
    ws.merge_cells('C8:E8')
    ws['C8'] = pesaje.get('choferNombre', '-')
    ws['C8'].font = normal_font
    
    ws.merge_cells('F8:G8')
    ws['F8'] = 'DNI:'
    ws['F8'].font = header_font
    ws['H8'] = pesaje.get('choferDni', '-')
    ws['H8'].font = normal_font
    
    # Fila 9: Patentes
    ws.merge_cells('A9:B9')
    ws['A9'] = 'Patente Chasis:'
    ws['A9'].font = header_font
    ws['C9'] = pesaje.get('patenteChasis', '-')
    ws['C9'].font = normal_font
    
    ws.merge_cells('D9:E9')
    ws['D9'] = 'Patente Acoplado:'
    ws['D9'].font = header_font
    ws['F9'] = pesaje.get('patenteAcoplado', '-')
    ws['F9'].font = normal_font
    
    # Fila 10: Documentos
    ws.merge_cells('A10:B10')
    ws['A10'] = 'DTE:'
    ws['A10'].font = header_font
    ws.merge_cells('C10:E10')
    ws['C10'] = data.get('dte', '-')
    ws['C10'].font = normal_font
    
    ws.merge_cells('F10:G10')
    ws['F10'] = 'Guía:'
    ws['F10'].font = header_font
    ws.merge_cells('H10:L10')
    ws['H10'] = data.get('guia', '-')
    ws['H10'].font = normal_font
    
    # Fila 11: Precintos
    ws.merge_cells('A11:B11')
    ws['A11'] = 'Precintos:'
    ws['A11'].font = header_font
    ws.merge_cells('C11:E11')
    ws['C11'] = pesaje.get('precintos', '-')
    ws['C11'].font = normal_font
    
    ws.merge_cells('F11:G11')
    ws['F11'] = 'Corral:'
    ws['F11'].font = header_font
    ws.merge_cells('H11:L11')
    ws['H11'] = data.get('corral', {}).get('nombre', '-') if data.get('corral') else '-'
    ws['H11'].font = normal_font
    
    # Fila 12: Espacio
    ws.row_dimensions[12].height = 15
    
    # Fila 13: Referencias de tipos
    ws.merge_cells('A13:B13')
    ws['A13'] = 'Referencias:'
    ws['A13'].font = header_font
    
    ws['C13'] = 'NOVILLITO'
    ws['C13'].font = small_font
    ws['D13'] = 'NT'
    ws['D13'].font = small_font
    
    ws['E13'] = 'NOVILLO'
    ws['E13'].font = small_font
    ws['F13'] = 'No'
    ws['F13'].font = small_font
    
    ws['G13'] = 'VAQUILLONA'
    ws['G13'].font = small_font
    ws['H13'] = 'Vq'
    ws['H13'].font = small_font
    
    ws['I13'] = 'VACA'
    ws['I13'].font = small_font
    ws['J13'] = 'Va'
    ws['J13'].font = small_font
    
    ws['K13'] = 'TORO'
    ws['K13'].font = small_font
    ws['L13'] = 'To'
    ws['L13'].font = small_font
    
    # Fila 14: Espacio
    ws.row_dimensions[14].height = 10
    
    # Fila 15: Encabezados de tabla
    headers = ['N°', 'TIPO ANIMAL', '', '', '', '', 'PESO ENTRADA', 'N° CORRAL', 'N° CARAVANA']
    header_positions = [(1, 'N°'), (2, 'TIPO DE ANIMAL'), (7, 'PESO ENTRADA'), (8, 'N° CORRAL'), (9, 'N° CARAVANA')]
    
    for col, header in header_positions:
        cell = ws.cell(row=15, column=col)
        cell.value = header
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border
        cell.fill = header_fill
    
    # Merge tipo animal
    ws.merge_cells('B15:F15')
    ws.merge_cells('A15:A16')
    ws.merge_cells('B15:F16')
    ws.merge_cells('G15:G16')
    ws.merge_cells('H15:H16')
    ws.merge_cells('I15:I16')
    
    # Fila 16: continuar encabezados
    for col in range(1, 10):
        ws.cell(row=16, column=col).border = thin_border
    
    # Datos de animales
    animales = data.get('animales', [])
    row_num = 17
    
    for i, animal in enumerate(animales[:50], 1):  # Máximo 50 por hoja
        # N°
        ws.cell(row=row_num, column=1).value = animal.get('numero', i)
        ws.cell(row=row_num, column=1).font = normal_font
        ws.cell(row=row_num, column=1).alignment = center_align
        ws.cell(row=row_num, column=1).border = thin_border
        
        # Tipo Animal (marcar con X en la columna correspondiente)
        tipo = animal.get('tipoAnimal', '')
        tipo_cols = {'NT': 2, 'NO': 3, 'VQ': 4, 'VA': 5, 'TO': 6}
        
        for col in range(2, 7):
            ws.cell(row=row_num, column=col).border = thin_border
            ws.cell(row=row_num, column=col).alignment = center_align
        
        if tipo in tipo_cols:
            ws.cell(row=row_num, column=tipo_cols[tipo]).value = 'X'
            ws.cell(row=row_num, column=tipo_cols[tipo]).font = normal_font
        
        # Peso
        peso = animal.get('pesoVivo')
        ws.cell(row=row_num, column=7).value = peso if peso else ''
        ws.cell(row=row_num, column=7).font = normal_font
        ws.cell(row=row_num, column=7).alignment = center_align
        ws.cell(row=row_num, column=7).border = thin_border
        
        # Corral
        ws.cell(row=row_num, column=8).value = data.get('corral', {}).get('nombre', '') if data.get('corral') else ''
        ws.cell(row=row_num, column=8).font = normal_font
        ws.cell(row=row_num, column=8).alignment = center_align
        ws.cell(row=row_num, column=8).border = thin_border
        
        # Caravana
        ws.cell(row=row_num, column=9).value = animal.get('caravana', '')
        ws.cell(row=row_num, column=9).font = normal_font
        ws.cell(row=row_num, column=9).alignment = center_align
        ws.cell(row=row_num, column=9).border = thin_border
        
        row_num += 1
    
    # Si hay más de 50 animales, agregar nota
    if len(animales) > 50:
        ws.cell(row=row_num + 1, column=1).value = f"... y {len(animales) - 50} animales más"
        ws.merge_cells(f'A{row_num + 1}:I{row_num + 1}')
    
    # Ajustar altura de filas
    for row in range(1, row_num + 2):
        ws.row_dimensions[row].height = 18
    
    # Guardar
    wb.save(output_path)
    return output_path

def main():
    if len(sys.argv) < 3:
        print(json.dumps({'error': 'Uso: script.py <data_json> <output_path>'}))
        sys.exit(1)
    
    try:
        data = json.loads(sys.argv[1])
        output_path = sys.argv[2]
        
        result = create_planilla_01(data, output_path)
        print(json.dumps({'success': True, 'path': result}))
    except Exception as e:
        print(json.dumps({'error': str(e)}))
        sys.exit(1)

if __name__ == '__main__':
    main()
