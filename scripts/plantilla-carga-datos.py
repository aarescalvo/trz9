import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

wb = Workbook()

# Colores y estilos
header_fill = PatternFill(start_color="1B3F66", end_color="1B3F66", fill_type="solid")
header_font = Font(name='Arial', color="FFFFFF", bold=True, size=10)
title_font = Font(name='Arial', size=14, bold=True)
input_font = Font(name='Arial', color="0000FF", size=10)  # Azul para inputs
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def create_sheet(wb, name, columns, examples=None, description=""):
    ws = wb.create_sheet(name)
    
    # Título
    ws['A1'] = name.upper()
    ws['A1'].font = title_font
    ws.merge_cells(f'A1:{chr(64+len(columns))}1')
    
    # Descripción
    if description:
        ws['A2'] = description
        ws['A2'].font = Font(italic=True, color="666666")
        ws.merge_cells(f'A2:{chr(64+len(columns))}2')
    
    # Headers
    header_row = 4
    for col, header in enumerate(columns, 1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = thin_border
        ws.column_dimensions[chr(64+col)].width = 15
    
    # Ejemplos
    if examples:
        for row_idx, example in enumerate(examples, header_row + 1):
            for col, value in enumerate(example, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.font = input_font
                cell.border = thin_border
    
    return ws

# ===== HOJA 1: INSTRUCCIONES =====
ws_inst = wb.active
ws_inst.title = "INSTRUCCIONES"

instructions = [
    ["PLANTILLA DE CARGA DE DATOS - TRAZASOLE v3.3.7"],
    [""],
    ["Esta plantilla contiene todas las hojas necesarias para cargar datos históricos."],
    [""],
    ["ORDEN DE CARGA (IMPORTANTE):"],
    ["1. PRODUCTORES - Datos de los productores/usuarios de faena"],
    ["2. CLIENTES - Clientes del sistema (si son diferentes a productores)"],
    ["3. CORRALES - Corrales disponibles"],
    ["4. TROPAS - Cada tropa con sus datos básicos"],
    ["5. ANIMALES - Animales de cada tropa (usar código de tropa)"],
    ["6. PESAJES_INDIVIDUALES - Pesos individuales de animales"],
    ["7. ASIGNACIONES_GARRONES - Asignación de garrones a animales"],
    ["8. ROMANEOS - Medias reses pesadas"],
    [""],
    ["CONVENCIONES:"],
    ["- Texto en AZUL = Datos que debe ingresar"],
    ["- Las columnas con * son obligatorias"],
    ["- Los códigos de tropa deben coincidir con la hoja TROPAS"],
    ["- Los números de animal deben coincidir con la hoja ANIMALES"],
    [""],
    ["TIPOS DE ANIMALES VALIDOS:"],
    ["BOVINO: TO (Toro), VA (Vaca), VQ (Vaquillona), MEJ (Torito/Mej), NO (Novillo), NT (Novillito)"],
    ["EQUINO: PADRILLO, POTRILLO, YEGUA, CABALLO, BURRO, MULA"],
    [""],
    ["RAZAS BOVINAS:"],
    ["Angus, Hereford, Braford, Brangus, Charolais, Limousin, Santa Gertrudis, Nelore, Brahman, Cebú, Cruza, Otro"],
    [""],
    ["ESTADOS DE ANIMAL:"],
    ["RECIBIDO, PESADO, EN_FAENA, FAENADO"],
    [""],
    ["CONTACTO: Si tiene dudas, consulte con el administrador del sistema"],
]

for row_idx, row in enumerate(instructions, 1):
    ws_inst.cell(row=row_idx, column=1, value=row[0] if row else "")
    if row_idx == 1:
        ws_inst.cell(row=row_idx, column=1).font = Font(size=16, bold=True)

ws_inst.column_dimensions['A'].width = 100

# ===== HOJA 2: PRODUCTORES =====
create_sheet(wb, "PRODUCTORES", 
    columns=["nombre*", "cuit", "direccion", "telefono", "email", "observaciones"],
    examples=[
        ["JUAN PEREZ", "20-12345678-9", "Ruta 9 Km 45", "351-123456", "juan@email.com", "Productor habitual"],
        ["MARIA GARCIA", "27-87654321-0", "Av. Principal 123", "358-987654", "", ""],
    ],
    description="Datos de los productores/usuarios de faena que envían animales"
)

# ===== HOJA 3: CLIENTES =====
create_sheet(wb, "CLIENTES",
    columns=["nombre*", "cuit", "direccion", "telefono", "email", "esProductor", "observaciones"],
    examples=[
        ["FRIGORIFICO SUR", "30-12345678-1", "Zona Industrial", "351-111111", "sur@frig.com", "NO", "Cliente habitual"],
        ["JUAN PEREZ", "20-12345678-9", "Ruta 9 Km 45", "351-123456", "", "SI", "Mismo que productor"],
    ],
    description="Clientes del sistema (pueden ser productores o compradores)"
)

# ===== HOJA 4: CORRALES =====
create_sheet(wb, "CORRALES",
    columns=["nombre*", "capacidad", "observaciones"],
    examples=[
        ["CORRAL 1", "50", "Corral principal"],
        ["CORRAL 2", "40", ""],
        ["CORRAL 3", "60", "Para equinos"],
    ],
    description="Corrales disponibles para alojar animales"
)

# ===== HOJA 5: TROPAS =====
create_sheet(wb, "TROPAS",
    columns=["codigo*", "fechaIngreso*", "especie*", "cantidadCabezas*", "productorNombre", "corralNombre", "pesoNeto", "estado", "observaciones"],
    examples=[
        ["B 2026 0001", "2026-01-15", "BOVINO", "10", "JUAN PEREZ", "CORRAL 1", "4500", "PESADO", "Primera tropa del año"],
        ["B 2026 0002", "2026-01-16", "BOVINO", "15", "MARIA GARCIA", "CORRAL 2", "6200", "PESADO", ""],
        ["E 2026 0001", "2026-01-20", "EQUINO", "5", "JUAN PEREZ", "CORRAL 3", "2500", "PESADO", "Equinos"],
    ],
    description="Tropas ingresadas. El código debe ser único. Formato: B/E AÑO NUMERO"
)

# ===== HOJA 6: ANIMALES =====
create_sheet(wb, "ANIMALES",
    columns=["tropaCodigo*", "numero*", "tipoAnimal*", "caravana", "raza", "pesoVivo", "estado", "observaciones"],
    examples=[
        ["B 2026 0001", "1", "VA", "1234", "Angus", "450", "PESADO", ""],
        ["B 2026 0001", "2", "VA", "1235", "Hereford", "420", "PESADO", ""],
        ["B 2026 0001", "3", "TO", "", "Braford", "580", "PESADO", "Sin caravana"],
        ["B 2026 0002", "1", "NO", "2001", "Angus", "480", "PESADO", ""],
    ],
    description="Animales de cada tropa. El número es secuencial dentro de la tropa."
)

# ===== HOJA 7: PESAJES INDIVIDUALES =====
create_sheet(wb, "PESAJES_INDIVIDUALES",
    columns=["tropaCodigo*", "numeroAnimal*", "peso*", "fecha", "observaciones"],
    examples=[
        ["B 2026 0001", "1", "450", "2026-01-15", ""],
        ["B 2026 0001", "2", "420", "2026-01-15", ""],
        ["B 2026 0001", "3", "580", "2026-01-15", ""],
    ],
    description="Pesajes individuales de cada animal. Debe coincidir con tropa y número de animal."
)

# ===== HOJA 8: ASIGNACIONES GARRONES =====
create_sheet(wb, "ASIGNACIONES_GARRONES",
    columns=["garron*", "tropaCodigo*", "numeroAnimal*", "fecha*", "tieneMediaDer", "tieneMediaIzq"],
    examples=[
        ["1", "B 2026 0001", "1", "2026-01-16", "SI", "SI"],
        ["2", "B 2026 0001", "2", "2026-01-16", "SI", "SI"],
        ["3", "B 2026 0001", "3", "2026-01-16", "SI", "SI"],
    ],
    description="Asignación de garrones a animales para faena. tieneMediaDer/Izq: SI o NO"
)

# ===== HOJA 9: ROMANEOS =====
create_sheet(wb, "ROMANEOS",
    columns=["garron*", "lado*", "peso*", "fecha*", "tropaCodigo", "denticion"],
    examples=[
        ["1", "DERECHA", "125.5", "2026-01-16", "B 2026 0001", "4"],
        ["1", "IZQUIERDA", "122.3", "2026-01-16", "B 2026 0001", "4"],
        ["2", "DERECHA", "118.0", "2026-01-16", "B 2026 0001", "6"],
        ["2", "IZQUIERDA", "115.5", "2026-01-16", "B 2026 0001", "6"],
    ],
    description="Pesaje de medias reses (romaneo). Lado: DERECHA o IZQUIERDA. Denticion: 0,2,4,6,8"
)

# Guardar
wb.save('/home/z/my-project/plantilla_carga_datos_trazasole.xlsx')
print("✅ Archivo creado: plantilla_carga_datos_trazasole.xlsx")
