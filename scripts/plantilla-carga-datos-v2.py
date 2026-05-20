"""
Plantilla de Carga de Datos - TRAZASOLE v3.7.24
Version 2.0 - Incluye todas las pestañas necesarias para una base de datos completa.

Pestañas existentes (SIN MODIFICAR):
  1. INSTRUCCIONES
  2. PRODUCTORES
  3. CLIENTES
  4. CORRALES
  5. TROPAS
  6. ANIMALES
  7. PESAJES_INDIVIDUALES
  8. ASIGNACIONES_GARRONES
  9. ROMANEOS

Pestañas NUEVAS agregadas:
  10. INSTRUCCIONES_COMPLEMENTO
  11. CAMARAS
  12. TIPIFICADORES
  13. TRANSPORTISTAS
  14. PRODUCTOS
  15. TIPOS_MENUDENCIA
  16. TIPOS_SERVICIO
  17. PESAJES_CAMION
  18. MEDIAS_RESES
  19. MENUDENCIAS
  20. DECOMISOS
  21. CUIEROS
  22. RENDERING
  23. DESPACHOS
  24. PRECIOS_SERVICIO
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

wb = Workbook()

# Colores y estilos
header_fill = PatternFill(start_color="1B3F66", end_color="1B3F66", fill_type="solid")
header_font = Font(name='Arial', color="FFFFFF", bold=True, size=10)
title_font = Font(name='Arial', size=14, bold=True)
input_font = Font(name='Arial', color="0000FF", size=10)  # Azul para inputs
new_header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")  # Verde para nuevas pestañas
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def create_sheet(wb, name, columns, examples=None, description="", is_new=False):
    ws = wb.create_sheet(name)
    
    fill = new_header_fill if is_new else header_fill
    
    # Titulo
    ws['A1'] = name.upper()
    ws['A1'].font = title_font
    ws.merge_cells(f'A1:{chr(64+len(columns))}1')
    
    # Descripcion
    if description:
        ws['A2'] = description
        ws['A2'].font = Font(italic=True, color="666666")
        ws.merge_cells(f'A2:{chr(64+len(columns))}2')
    
    # Headers
    header_row = 4
    for col, header in enumerate(columns, 1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.fill = fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = thin_border
        ws.column_dimensions[chr(64+col)].width = 18
    
    # Ejemplos
    if examples:
        for row_idx, example in enumerate(examples, header_row + 1):
            for col, value in enumerate(example, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.font = input_font
                cell.border = thin_border
    
    return ws


# ============================================================================
# PESTANAS EXISTENTES (SIN MODIFICAR)
# ============================================================================

# ===== HOJA 1: INSTRUCCIONES =====
ws_inst = wb.active
ws_inst.title = "INSTRUCCIONES"

instructions = [
    ["PLANTILLA DE CARGA DE DATOS - TRAZASOLE v3.3.7"],
    [""],
    ["Esta plantilla contiene todas las hojas necesarias para cargar datos historicos."],
    [""],
    ["ORDEN DE CARGA (IMPORTANTE):"],
    ["1. PRODUCTORES - Datos de los productores/usuarios de faena"],
    ["2. CLIENTES - Clientes del sistema (si son diferentes a productores)"],
    ["3. CORRALES - Corrales disponibles"],
    ["4. TROPAS - Cada tropa con sus datos basicos"],
    ["5. ANIMALES - Animales de cada tropa (usar codigo de tropa)"],
    ["6. PESAJES_INDIVIDUALES - Pesos individuales de animales"],
    ["7. ASIGNACIONES_GARRONES - Asignacion de garrones a animales"],
    ["8. ROMANEOS - Medias reses pesadas"],
    [""],
    ["CONVENCIONES:"],
    ["- Texto en AZUL = Datos que debe ingresar"],
    ["- Las columnas con * son obligatorias"],
    ["- Los codigos de tropa deben coincidir con la hoja TROPAS"],
    ["- Los numeros de animal deben coincidir con la hoja ANIMALES"],
    [""],
    ["TIPOS DE ANIMALES VALIDOS:"],
    ["BOVINO: TO (Toro), VA (Vaca), VQ (Vaquillona), MEJ (Torito/Mej), NO (Novillo), NT (Novillito)"],
    ["EQUINO: PADRILLO, POTRILLO, YEGUA, CABALLO, BURRO, MULA"],
    [""],
    ["RAZAS BOVINAS:"],
    ["Angus, Hereford, Braford, Brangus, Charolais, Limousin, Santa Gertrudis, Nelore, Brahman, Cebu, Cruza, Otro"],
    [""],
    ["ESTADOS DE ANIMAL:"],
    ["RECIBIDO, PESADO, EN_FAENA, FAENADO"],
    [""],
    ["CONTACTO: Si tiene dudas, consulte con el administrador del sistema"],
]

for row_idx, row in enumerate(instructions, 1):
    ws_inst.cell(row=row_idx, column=1, value=row[0] if row else "")
    if row_idx == 1:
        ws_inst.cell(row=row_idx, column=1).font = Font(size=16, bold=True)

ws_inst.column_dimensions['A'].width = 100

# ===== HOJA 2: PRODUCTORES =====
create_sheet(wb, "PRODUCTORES", 
    columns=["nombre*", "cuit", "direccion", "telefono", "email", "observaciones"],
    examples=[
        ["JUAN PEREZ", "20-12345678-9", "Ruta 9 Km 45", "351-123456", "juan@email.com", "Productor habitual"],
        ["MARIA GARCIA", "27-87654321-0", "Av. Principal 123", "358-987654", "", ""],
    ],
    description="Datos de los productores/usuarios de faena que envian animales"
)

# ===== HOJA 3: CLIENTES =====
create_sheet(wb, "CLIENTES",
    columns=["nombre*", "cuit", "direccion", "telefono", "email", "esProductor", "observaciones"],
    examples=[
        ["FRIGORIFICO SUR", "30-12345678-1", "Zona Industrial", "351-111111", "sur@frig.com", "NO", "Cliente habitual"],
        ["JUAN PEREZ", "20-12345678-9", "Ruta 9 Km 45", "351-123456", "", "SI", "Mismo que productor"],
    ],
    description="Clientes del sistema (pueden ser productores o compradores)"
)

# ===== HOJA 4: CORRALES =====
create_sheet(wb, "CORRALES",
    columns=["nombre*", "capacidad", "observaciones"],
    examples=[
        ["CORRAL 1", "50", "Corral principal"],
        ["CORRAL 2", "40", ""],
        ["CORRAL 3", "60", "Para equinos"],
    ],
    description="Corrales disponibles para alojar animales"
)

# ===== HOJA 5: TROPAS =====
create_sheet(wb, "TROPAS",
    columns=["codigo*", "fechaIngreso*", "especie*", "cantidadCabezas*", "productorNombre", "corralNombre", "pesoNeto", "estado", "observaciones"],
    examples=[
        ["B 2026 0001", "2026-01-15", "BOVINO", "10", "JUAN PEREZ", "CORRAL 1", "4500", "PESADO", "Primera tropa del ano"],
        ["B 2026 0002", "2026-01-16", "BOVINO", "15", "MARIA GARCIA", "CORRAL 2", "6200", "PESADO", ""],
        ["E 2026 0001", "2026-01-20", "EQUINO", "5", "JUAN PEREZ", "CORRAL 3", "2500", "PESADO", "Equinos"],
    ],
    description="Tropas ingresadas. El codigo debe ser unico. Formato: B/E ANO NUMERO"
)

# ===== HOJA 6: ANIMALES =====
create_sheet(wb, "ANIMALES",
    columns=["tropaCodigo*", "numero*", "tipoAnimal*", "caravana", "raza", "pesoVivo", "estado", "observaciones"],
    examples=[
        ["B 2026 0001", "1", "VA", "1234", "Angus", "450", "PESADO", ""],
        ["B 2026 0001", "2", "VA", "1235", "Hereford", "420", "PESADO", ""],
        ["B 2026 0001", "3", "TO", "", "Braford", "580", "PESADO", "Sin caravana"],
        ["B 2026 0002", "1", "NO", "2001", "Angus", "480", "PESADO", ""],
    ],
    description="Animales de cada tropa. El numero es secuencial dentro de la tropa."
)

# ===== HOJA 7: PESAJES INDIVIDUALES =====
create_sheet(wb, "PESAJES_INDIVIDUALES",
    columns=["tropaCodigo*", "numeroAnimal*", "peso*", "fecha", "observaciones"],
    examples=[
        ["B 2026 0001", "1", "450", "2026-01-15", ""],
        ["B 2026 0001", "2", "420", "2026-01-15", ""],
        ["B 2026 0001", "3", "580", "2026-01-15", ""],
    ],
    description="Pesajes individuales de cada animal. Debe coincidir con tropa y numero de animal."
)

# ===== HOJA 8: ASIGNACIONES GARRONES =====
create_sheet(wb, "ASIGNACIONES_GARRONES",
    columns=["garron*", "tropaCodigo*", "numeroAnimal*", "fecha*", "tieneMediaDer", "tieneMediaIzq"],
    examples=[
        ["1", "B 2026 0001", "1", "2026-01-16", "SI", "SI"],
        ["2", "B 2026 0001", "2", "2026-01-16", "SI", "SI"],
        ["3", "B 2026 0001", "3", "2026-01-16", "SI", "SI"],
    ],
    description="Asignacion de garrones a animales para faena. tieneMediaDer/Izq: SI o NO"
)

# ===== HOJA 9: ROMANEOS =====
create_sheet(wb, "ROMANEOS",
    columns=["garron*", "lado*", "peso*", "fecha*", "tropaCodigo", "denticion"],
    examples=[
        ["1", "DERECHA", "125.5", "2026-01-16", "B 2026 0001", "4"],
        ["1", "IZQUIERDA", "122.3", "2026-01-16", "B 2026 0001", "4"],
        ["2", "DERECHA", "118.0", "2026-01-16", "B 2026 0001", "6"],
        ["2", "IZQUIERDA", "115.5", "2026-01-16", "B 2026 0001", "6"],
    ],
    description="Pesaje de medias reses (romaneo). Lado: DERECHA o IZQUIERDA. Denticion: 0,2,4,6,8"
)


# ============================================================================
# PESTANAS NUEVAS
# ============================================================================

# ===== HOJA 10: INSTRUCCIONES COMPLEMENTO =====
ws_inst2 = wb.create_sheet("INSTRUCCIONES_COMPLEMENTO")

new_instructions = [
    ["PLANTILLA DE CARGA DE DATOS - COMPLEMENTO v2.0"],
    [""],
    ["Las siguientes pestanas se agregaron para tener una base de datos completa y actualizada."],
    ["Los headers en VERDE identifican las nuevas pestanas."],
    [""],
    ["ORDEN DE CARGA DE NUEVAS PESTANAS (despues de las existentes):"],
    ["  A) DATOS MAESTROS (cargar primero, son requisito para los operativos):"],
    ["     10. CAMARAS - Camaras frigorificas donde se alojan las medias reses"],
    ["     11. TIPIFICADORES - Personal que tipifica/clasifica la carne"],
    ["     12. TRANSPORTISTAS - Empresas de transporte"],
    ["     13. PRODUCTOS - Catalogo de productos (para despostada, menudencias, empaques)"],
    ["     14. TIPOS_MENUDENCIA - Categorias de menudencias"],
    ["     15. TIPOS_SERVICIO - Tipos de servicio que se facturan"],
    [""],
    ["  B) DATOS OPERATIVOS (cargar despues de los maestros):"],
    ["     16. PESAJES_CAMION - Pesaje de camiones (ingreso hacienda / salida mercaderia)"],
    ["     17. MEDIAS_RESES - Registro individual de medias reses despues del romaneo"],
    ["     18. MENUDENCIAS - Registro de menudencias por tropa"],
    ["     19. DECOMISOS - Decomisos totales o parciales"],
    ["     20. CUIEROS - Registro de cueros"],
    ["     21. RENDERING - Registro de grasa, desperdicios, fondo digestor, sangre"],
    ["     22. DESPACHOS - Despachos/expediciones de mercaderia"],
    ["     23. PRECIOS_SERVICIO - Precios de servicios por cliente"],
    [""],
    ["CONVENCIONES ADICIONALES:"],
    ["- Los campos que referencian a otra pestana (ej: camaraNombre, tipificadorMatricula)"],
    ["  deben coincidir EXACTAMENTE con el nombre/matricula de la pestana referenciada"],
    ["- Las fechas en formato YYYY-MM-DD (ej: 2026-01-15)"],
    ["- Los campos con * son obligatorios"],
    ["- Los campos sin * son opcionales"],
    ["- NO repetir datos ya cargados en otras pestanas (usar referencias)"],
    [""],
    ["TIPOS DE CAMARA: FAENA (capacidad en ganchos), CUARTEO (capacidad en KG), DEPOSITO (capacidad en KG)"],
    ["SIGLAS DE MEDIA RES: A (Asado), T (Trasero), D (Delantero)"],
    ["ESTADOS DE MEDIA RES: EN_CAMARA, EN_CUARTEO, DESPACHADO"],
    ["CONSERVACION DE CUERO: SALADO, FRESCO, CORTADO"],
    ["TIPOS DE RENDERING: GRASA, DESPERDICIOS, FONDO_DIGESTOR, SANGRE"],
    ["METODOS DE PAGO: EFECTIVO, TRANSFERENCIA, CHEQUE, TARJETA"],
    ["CONDICIONES IVA: RI (Responsable Inscripto), CF (Consumidor Final), MT (Monotributo), EX (Exento)"],
]

for row_idx, row in enumerate(new_instructions, 1):
    ws_inst2.cell(row=row_idx, column=1, value=row[0] if row else "")
    if row_idx == 1:
        ws_inst2.cell(row=row_idx, column=1).font = Font(size=16, bold=True)

ws_inst2.column_dimensions['A'].width = 110


# ===== HOJA 11: CAMARAS =====
create_sheet(wb, "CAMARAS",
    columns=["nombre*", "tipo*", "capacidad", "observaciones"],
    examples=[
        ["CAMARA 1", "FAENA", "200", "Camara principal de faena"],
        ["CAMARA 2", "FAENA", "150", "Camara secundaria"],
        ["CAMARA 3", "CUARTEO", "5000", "Camara de cuarteo"],
        ["CAMARA 4", "DEPOSITO", "10000", "Deposito de mercaderia"],
    ],
    description="Camaras frigorificas. Tipo: FAENA (ganchos), CUARTEO (KG), DEPOSITO (KG). Requerido antes de cargar medias reses.",
    is_new=True
)

# ===== HOJA 12: TIPIFICADORES =====
create_sheet(wb, "TIPIFICADORES",
    columns=["nombre*", "apellido*", "matricula*", "numeroInterno", "observaciones"],
    examples=[
        ["Carlos", "Rodriguez", "TIP-001", "1", "Tipificador principal"],
        ["Maria", "Lopez", "TIP-002", "2", ""],
    ],
    description="Tipificadores que clasifican la carne en el romaneo. La matricula es unica y se usa en los romaneos.",
    is_new=True
)

# ===== HOJA 13: TRANSPORTISTAS =====
create_sheet(wb, "TRANSPORTISTAS",
    columns=["nombre*", "cuit", "direccion", "telefono", "observaciones"],
    examples=[
        ["TRANSPORTES DEL SUR SRL", "30-55555555-1", "Ruta 9 Km 10", "351-555555", "Transportista habitual"],
        ["LOGISTICA GARCIA", "20-44444444-2", "Av. Industrial 456", "358-444444", ""],
    ],
    description="Empresas de transporte que llevan hacienda y/o mercaderia.",
    is_new=True
)

# ===== HOJA 14: PRODUCTOS =====
create_sheet(wb, "PRODUCTOS",
    columns=["codigo*", "nombre*", "especie*", "codigoTipificacion", "codigoTipoTrabajo", "codigoTransporte", "codigoDestino", "tara", "diasConservacion", "requiereTipificacion", "tipoRotulo", "precio", "apareceRendimiento", "apareceStock", "observaciones"],
    examples=[
        ["001", "MEDIA RES BOVINA", "BOVINO", ".00", "0", "0", ".00", "0", "180", "NO", "MEDIA_RES", "", "SI", "SI", "Producto principal"],
        ["002", "CUARTO DELANTERO", "BOVINO", ".01", "0", "0", ".00", "0", "180", "NO", "CUARTO", "", "SI", "SI", ""],
        ["003", "CUARTO TRASERO", "BOVINO", ".02", "0", "0", ".00", "0", "180", "NO", "CUARTO", "", "SI", "SI", ""],
        ["010", "BIFE ANGOSTO", "BOVINO", ".10", "1", "0", ".00", "0.2", "90", "SI", "PRODUCTO_TERMINADO_ENVASE_PRIMARIO", "25000", "NO", "SI", "Corte premium"],
        ["050", "CORAZON", "BOVINO", "", "0", "0", ".00", "0", "30", "NO", "MENUDENCIA", "8000", "SI", "SI", "Menudencia"],
    ],
    description="Catalogo de productos. Codigo: 3 digitos. Especie: BOVINO o EQUINO. TipoRotulo: MEDIA_RES, CUARTO, MENUDENCIA, PRODUCTO_TERMINADO_ENVASE_PRIMARIO, etc.",
    is_new=True
)

# ===== HOJA 15: TIPOS MENUDENCIA =====
create_sheet(wb, "TIPOS_MENUDENCIA",
    columns=["nombre*", "observaciones"],
    examples=[
        ["CORAZON", "Menudencia roja"],
        ["HIGADO", "Menudencia roja"],
        ["LENGUA", "Menudencia roja"],
        ["MONDONGO", "Menudencia blanca"],
        ["RIÑON", "Menudencia roja"],
        ["SESOS", "Menudencia roja"],
        ["MOLLEJAS", "Menudencia blanca"],
        ["CARRIL", "Menudencia roja"],
        ["PARRILLA", "Menudencia blanca"],
        ["RABO", "Menudencia roja"],
    ],
    description="Categorias de menudencias. Se usan para clasificar las menudencias por tropa.",
    is_new=True
)

# ===== HOJA 16: TIPOS SERVICIO =====
create_sheet(wb, "TIPOS_SERVICIO",
    columns=["codigo*", "nombre*", "descripcion", "unidad", "porcentajeIva", "seFactura", "observaciones"],
    examples=[
        ["FAENA", "Servicio de Faena x Kg", "Faena de animales por kilogramo de gancho", "KG", "21", "SI", "Servicio principal"],
        ["FAENA_CABEZA", "Servicio de Faena x Cabeza", "Faena por cabeza de animal", "UN", "21", "SI", "Tarifa fija por animal"],
        ["EMBOLSADO_VACIO", "Embolse al Vacio", "Embalaje al vacio de medias reses", "KG", "21", "SI", "Servicio adicional"],
        ["DESHUESADO", "Deshuesado / Despostada", "Proceso de deshuesado de medias", "KG", "21", "SI", "Para despostada"],
        ["CUARTEO", "Cuarteo", "Corte de media en delantero y trasero", "UN", "21", "SI", "Por media res"],
        ["TASA_INSPECCION", "Tasa Inspeccion Veterinaria", "Tasa obligatoria de inspeccion", "UN", "0", "SI", "Tasa fija por cabeza"],
        ["ARANCEL_IPCVA", "Arancel IPCVA", "Arancel del Instituto de Promocion de la Carne", "UN", "0", "SI", "Arancel fijo por cabeza"],
    ],
    description="Tipos de servicio que se facturan a los clientes. Codigo unico. Unidad: KG, UN, HORA.",
    is_new=True
)

# ===== HOJA 17: PESAJES CAMION =====
create_sheet(wb, "PESAJES_CAMION",
    columns=["tipo*", "numeroTicket*", "patenteChasis*", "patenteAcoplado", "choferNombre", "choferDni", "transportistaNombre", "tropaCodigo", "destino", "remito", "pesoBruto", "pesoTara", "pesoNeto", "precintos", "fecha", "estado", "observaciones"],
    examples=[
        ["INGRESO_HACIENDA", "1001", "AB123CD", "EF456GH", "Jose Martinez", "12345678", "TRANSPORTES DEL SUR SRL", "B 2026 0001", "", "", "28000", "8000", "20000", "", "2026-01-15", "CERRADO", "Ingreso tropa 1"],
        ["INGRESO_HACIENDA", "1002", "CD789EF", "", "Pedro Gomez", "87654321", "LOGISTICA GARCIA", "B 2026 0002", "", "", "32000", "8500", "23500", "", "2026-01-16", "CERRADO", ""],
        ["SALIDA_MERCADERIA", "2001", "GH012IJ", "KL345MN", "Luis Sanchez", "11222333", "TRANSPORTES DEL SUR SRL", "", "Supermercado del Centro", "R-0001", "15000", "5000", "10000", "P001,P002", "2026-01-20", "CERRADO", "Despacho 1"],
        ["PESAJE_PARTICULAR", "3001", "MN678OP", "", "Cliente particular", "", "", "", "", "", "5000", "2000", "3000", "", "2026-01-22", "CERRADO", "Pesaje particular"],
    ],
    description="Pesajes de camiones. Tipo: INGRESO_HACIENDA, SALIDA_MERCADERIA, PESAJE_PARTICULAR. Para INGRESO indicar tropaCodigo. Para SALIDA indicar destino y remito.",
    is_new=True
)

# ===== HOJA 18: MEDIAS RESES (ROMANEO DE LO FAENADO) =====
create_sheet(wb, "MEDIAS_RESES",
    columns=["garron*", "lado*", "peso*", "sigla*", "camaraNombre*", "usuarioFaenaNombre", "fecha*", "estado", "observaciones"],
    examples=[
        ["1", "IZQUIERDA", "125.5", "A", "CAMARA 1", "FRIGORIFICO SUR", "2026-01-16", "EN_CAMARA", ""],
        ["1", "DERECHA", "122.3", "A", "CAMARA 1", "FRIGORIFICO SUR", "2026-01-16", "EN_CAMARA", ""],
        ["2", "IZQUIERDA", "118.0", "A", "CAMARA 1", "FRIGORIFICO SUR", "2026-01-16", "EN_CAMARA", ""],
        ["2", "DERECHA", "115.5", "A", "CAMARA 1", "FRIGORIFICO SUR", "2026-01-16", "EN_CAMARA", ""],
        ["3", "IZQUIERDA", "130.2", "A", "CAMARA 2", "LA AZUL SELECCION", "2026-01-16", "EN_CAMARA", ""],
        ["3", "DERECHA", "128.8", "A", "CAMARA 2", "LA AZUL SELECCION", "2026-01-16", "EN_CAMARA", ""],
    ],
    description="ROMANEO DE LO FAENADO: Registro individual de cada media res despues del romaneo. Indica en que camara se ubica y a que usuario pertenece. Lado: IZQUIERDA o DERECHA. Sigla: A (Asado), T (Trasero), D (Delantero). Estado: EN_CAMARA, EN_CUARTEO, DESPACHADO.",
    is_new=True
)

# ===== HOJA 19: MENUDENCIAS =====
create_sheet(wb, "MENUDENCIAS",
    columns=["tipoMenudenciaNombre*", "tropaCodigo*", "pesoIngreso", "pesoElaborado", "cantidadBolsas", "numeroBolsa", "fechaIngreso*", "fechaElaboracion", "observaciones"],
    examples=[
        ["CORAZON", "B 2026 0001", "15.5", "14.8", "3", "1", "2026-01-16", "2026-01-16", "3 corazones de la tropa"],
        ["HIGADO", "B 2026 0001", "22.0", "21.0", "5", "2", "2026-01-16", "2026-01-16", ""],
        ["LENGUA", "B 2026 0001", "8.5", "8.2", "2", "3", "2026-01-16", "2026-01-16", ""],
        ["MONDONGO", "B 2026 0001", "18.0", "17.5", "4", "4", "2026-01-16", "2026-01-17", "Se elabora al dia siguiente"],
        ["CORAZON", "B 2026 0002", "12.0", "11.5", "2", "5", "2026-01-17", "2026-01-17", ""],
    ],
    description="Registro de menudencias por tropa. tipoMenudenciaNombre debe coincidir con TIPOS_MENUDENCIA. tropaCodigo con TROPAS.",
    is_new=True
)

# ===== HOJA 20: DECOMISOS =====
create_sheet(wb, "DECOMISOS",
    columns=["garron*", "tipo*", "peso*", "motivo", "tropaCodigo", "fecha*", "observaciones"],
    examples=[
        ["5", "PARCIAL", "8.5", "Lesion hepatica", "B 2026 0001", "2026-01-16", "Decomiso parcial de higado"],
        ["8", "TOTAL", "135.0", "Tuberculosis", "B 2026 0002", "2026-01-17", "Decomiso total del animal"],
        ["12", "PARCIAL", "3.2", "Absceso", "B 2026 0002", "2026-01-17", "Decomiso parcial de rinon"],
    ],
    description="Decomisos de la faena. Tipo: TOTAL (animal completo) o PARCIAL (parte del animal). Relacionado con garron y tropaCodigo.",
    is_new=True
)

# ===== HOJA 21: CUIEROS =====
create_sheet(wb, "CUIEROS",
    columns=["tropaCodigo*", "cantidad*", "pesoKg*", "conservacion*", "destino", "tipoDestino", "remito", "fechaDespacho", "observaciones"],
    examples=[
        ["B 2026 0001", "10", "450.0", "SALADO", "Curtiembre San Martin", "CURTIEMBRE", "C-0001", "2026-01-20", "10 cueros de la tropa"],
        ["B 2026 0002", "15", "680.0", "SALADO", "Curtiembre San Martin", "CURTIEMBRE", "C-0002", "2026-01-22", ""],
        ["B 2026 0003", "5", "220.0", "FRESCO", "Venta directa a cliente", "VENTA_DIRECTA", "", "", "Venta sin despacho aun"],
    ],
    description="Registro de cueros por tropa. Conservacion: SALADO, FRESCO, CORTADO. TipoDestino: CURTIEMBRE o VENTA_DIRECTA.",
    is_new=True
)

# ===== HOJA 22: RENDERING =====
create_sheet(wb, "RENDERING",
    columns=["tipo*", "tropaCodigo", "pesoKg*", "destino", "fechaFaena*", "remito", "fechaDespacho", "observaciones"],
    examples=[
        ["GRASA", "B 2026 0001", "85.0", "Rendering SA", "2026-01-16", "R-REND-001", "2026-01-20", "Grasa de la tropa"],
        ["DESPERDICIOS", "B 2026 0001", "45.0", "Rendering SA", "2026-01-16", "R-REND-001", "2026-01-20", ""],
        ["SANGRE", "B 2026 0002", "120.0", "Rendering SA", "2026-01-17", "", "", "Aun no despachada"],
        ["FONDO_DIGESTOR", "B 2026 0002", "30.0", "Rendering SA", "2026-01-17", "", "", ""],
    ],
    description="Registro de rendering (subproductos de faena). Tipo: GRASA, DESPERDICIOS, FONDO_DIGESTOR, SANGRE.",
    is_new=True
)

# ===== HOJA 23: DESPACHOS =====
create_sheet(wb, "DESPACHOS",
    columns=["numero*", "fecha*", "destino*", "direccionDestino", "patenteCamion", "patenteAcoplado", "chofer", "choferDni", "transportistaNombre", "remito", "kgTotal", "cantidadMedias", "precintos", "estado", "observaciones"],
    examples=[
        ["1", "2026-01-20", "Supermercado del Centro", "Av. Libertador 800", "GH012IJ", "KL345MN", "Luis Sanchez", "11222333", "TRANSPORTES DEL SUR SRL", "R-0001", "244.0", "2", "P001,P002", "DESPACHADO", "2 medias de tropa B 2026 0001"],
        ["2", "2026-01-22", "Carniceria Don Jose", "San Martin 250", "QR789ST", "", "Ana Torres", "44555666", "LOGISTICA GARCIA", "R-0002", "487.0", "4", "P003,P004", "DESPACHADO", "4 medias de tropa B 2026 0002"],
    ],
    description="Despachos/expediciones de mercaderia. Numero correlativo. Los items (medias reses) se asocian en el sistema. Estado: PENDIENTE, EN_CARGA, DESPACHADO, ENTREGADO, ANULADO.",
    is_new=True
)

# ===== HOJA 24: PRECIOS SERVICIO =====
create_sheet(wb, "PRECIOS_SERVICIO",
    columns=["clienteNombre*", "tipoServicioCodigo*", "precio*", "fechaDesde*", "fechaHasta", "observaciones"],
    examples=[
        ["FRIGORIFICO SUR", "FAENA", "3500", "2026-01-01", "", "Precio vigente 2026"],
        ["FRIGORIFICO SUR", "TASA_INSPECCION", "500", "2026-01-01", "", "Tasa fija por cabeza"],
        ["FRIGORIFICO SUR", "ARANCEL_IPCVA", "300", "2026-01-01", "", "Arancel fijo por cabeza"],
        ["LA AZUL SELECCION", "FAENA", "3800", "2026-01-01", "", "Precio diferenciado"],
        ["LA AZUL SELECCION", "EMBOLSADO_VACIO", "1500", "2026-01-01", "", "Servicio adicional"],
        ["CARNICERIA DON JOSE", "FAENA", "3500", "2026-01-01", "", "Precio estandar"],
        ["CARNICERIA DON JOSE", "CUARTEO", "800", "2026-01-01", "", "Por media res"],
    ],
    description="Precios de servicios por cliente. clienteNombre debe coincidir con CLIENTES. tipoServicioCodigo con TIPOS_SERVICIO. fechaHasta vacio = vigente actualmente.",
    is_new=True
)


# Guardar
output_path = '/home/z/my-project/download/PLANTILLA_CARGA_DATOS_COMPLETA_v2.xlsx'
wb.save(output_path)
print(f"Archivo creado: {output_path}")
print(f"Total pestanas: {len(wb.sheetnames)}")
print(f"Pestanas: {', '.join(wb.sheetnames)}")
